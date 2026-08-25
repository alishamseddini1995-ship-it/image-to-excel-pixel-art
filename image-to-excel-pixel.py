import openpyxl
from PIL import Image
from openpyxl.styles import PatternFill

img = Image.open(r"C:\Users\ALI\OneDrive\Desktop\python\inception.jfif")

colorful_excel = openpyxl.Workbook()
ws = colorful_excel.active
width, height = img.size
for y in range(0, height, 10):
    for x in range(0, width, 10):

        block = img.crop((x, y, x + 10, y + 10))

        pixel_colors = block.getdata()

        r_total = 0
        g_total = 0
        b_total = 0

        for pixel in pixel_colors:
            r, g, b = pixel

            r_total += r
            g_total += g
            b_total += b

        r_average = r_total / len(pixel_colors)
        g_average = g_total / len(pixel_colors)
        b_average = b_total / len(pixel_colors)

# cell identification
        column = x // 10 + 1
        row = y // 10 + 1

        cell = ws.cell(row=row, column=column)

#  RGB to Hex
        hex_color = f"{int(r_average):02X}{int(g_average):02X}{int(b_average):02X}"

# cell coloring
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=hex_color
        )

# save file
colorful_excel.save("pixel_image.xlsx")